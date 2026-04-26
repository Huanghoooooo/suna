#!/usr/bin/env python3
"""
Portable invoice generator for bundled XLSX templates.

The templates are not fully uniform. This generator discovers the target cells
from labels in each sheet instead of relying on fixed coordinates.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


def _configure_text_output() -> None:
    requested = (os.getenv("PYTHONIOENCODING") or "").split(":", 1)[0] or None
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure:
            reconfigure(
                encoding=requested or getattr(stream, "encoding", None) or "utf-8",
                errors="replace",
            )


_configure_text_output()

SKILL_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = SKILL_DIR / "templates"
OUTPUT_DIR = SKILL_DIR / "output"

COUNTRY_SHEET = {
    "德国": "德国",
    "germany": "德国",
    "de": "德国",
    "英国": "英国",
    "uk": "英国",
    "gb": "英国",
    "united kingdom": "英国",
    "法国": "法国",
    "france": "法国",
    "fr": "法国",
    "意大利": "意大利",
    "italy": "意大利",
    "it": "意大利",
    "西班牙": "西班牙",
    "spain": "西班牙",
    "es": "西班牙",
    "波兰": "波兰",
    "poland": "波兰",
    "pl": "波兰",
    "德语发票": "德语发票",
    "german invoice": "德语发票",
}
DISPLAY_COUNTRIES = ["德国", "英国", "法国", "意大利", "西班牙", "波兰", "德语发票"]
NON_BUSINESS_SHEETS = {"税率计算", "社区企业不用征税"}
STORE_TEMPLATES: dict[str, Path] = {}


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def safe_filename(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*\s]+', "-", str(value)).strip("-") or "invoice"


def cell_coord(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


class InvoiceGenerator:
    def __init__(self) -> None:
        self.templates_dir = TEMPLATES_DIR
        self.output_dir = OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._scan_templates()

    def _scan_templates(self) -> None:
        STORE_TEMPLATES.clear()
        for file_path in sorted(self.templates_dir.glob("*.xlsx")):
            name = file_path.stem.replace("invoice 模板-", "").replace("invoice模板-", "")
            if name and name != file_path.stem:
                STORE_TEMPLATES[name] = file_path

    def list_stores(self) -> list[str]:
        return sorted(STORE_TEMPLATES.keys())

    def list_countries(self) -> list[str]:
        return DISPLAY_COUNTRIES

    def list_sheets(self, store_name: str) -> list[str]:
        template = self.find_template(store_name)
        wb = openpyxl.load_workbook(template, read_only=True, data_only=False)
        try:
            return [name for name in wb.sheetnames if name not in NON_BUSINESS_SHEETS]
        finally:
            wb.close()

    def find_template(self, store_name: str) -> Path:
        if store_name in STORE_TEMPLATES:
            return STORE_TEMPLATES[store_name]
        for name, file_path in STORE_TEMPLATES.items():
            if store_name in name or name in store_name:
                return file_path
        raise FileNotFoundError(
            f"No template found for store '{store_name}'. Available: {self.list_stores()}"
        )

    def find_sheet(
        self,
        wb: openpyxl.Workbook,
        *,
        sheet: str | None = None,
        country: str | None = None,
        credit_note: bool = False,
    ) -> str:
        if credit_note:
            target = "贷记单"
        elif sheet:
            target = sheet
        elif country:
            target = COUNTRY_SHEET.get(country.lower(), COUNTRY_SHEET.get(country, country))
        else:
            raise ValueError("Either sheet, country, or credit_note is required.")

        if target in wb.sheetnames:
            return target

        target_norm = norm(target)
        for name in wb.sheetnames:
            if norm(name) == target_norm:
                return name
        for name in wb.sheetnames:
            if target_norm in norm(name) or norm(name) in target_norm:
                return name
        raise ValueError(f"Sheet '{target}' not found. Available: {wb.sheetnames}")

    def generate(self, **kwargs: Any) -> Path:
        store = kwargs["store"]
        country = kwargs.get("country") or ""
        sheet = kwargs.get("sheet") or ""
        order_number = kwargs["order_number"]
        invoice_number = kwargs.get("invoice_number") or order_number
        credit_note = bool(kwargs.get("credit_note", False))
        currency = kwargs.get("currency", "EUR")
        invoice_date = self._parse_date(kwargs["invoice_date"])
        delivery_date = self._parse_date(kwargs["delivery_date"])
        customer_info = kwargs["customer_info"]
        product_description = kwargs["product_description"]
        quantity = int(kwargs["quantity"])
        unit_price = float(kwargs["unit_price"])

        template_path = self.find_template(store)
        output_sheet = "贷记单" if credit_note else sheet or country
        output_xlsx = (
            self.output_dir
            / f"{safe_filename(store)}_{safe_filename(output_sheet)}_{safe_filename(order_number)}_{'credit_note' if credit_note else 'invoice'}.xlsx"
        )
        shutil.copy(template_path, output_xlsx)

        wb = openpyxl.load_workbook(output_xlsx)
        sheet_name = self.find_sheet(
            wb, sheet=sheet or None, country=country or None, credit_note=credit_note
        )
        ws = wb[sheet_name]
        self._fill_sheet(
            ws,
            template_path=template_path,
            sheet_name=sheet_name,
            invoice_date=invoice_date,
            customer_info=customer_info,
            order_number=order_number,
            invoice_number=invoice_number,
            delivery_date=delivery_date,
            product_description=product_description,
            quantity=quantity,
            unit_price=-abs(unit_price) if credit_note else unit_price,
            currency=currency,
        )
        wb.save(output_xlsx)
        wb.close()
        return output_xlsx

    def _fill_sheet(
        self,
        ws,
        *,
        template_path: Path,
        sheet_name: str,
        invoice_date: datetime,
        customer_info: str,
        order_number: str,
        invoice_number: str,
        delivery_date: datetime,
        product_description: str,
        quantity: int,
        unit_price: float,
        currency: str,
    ) -> None:
        layout = self._discover_layout(ws, template_path=template_path, sheet_name=sheet_name)

        self._safe_set(ws, layout["invoice_date"], invoice_date)
        self._safe_set(ws, layout["customer_info"], customer_info)
        self._safe_set(ws, layout["order_number"], f"# {order_number}")
        self._safe_set(ws, layout["invoice_number"], invoice_number)
        self._safe_set(ws, layout["delivery_date"], delivery_date)

        product = layout["product"]
        product_row = product["row"]
        total_coord = cell_coord(product_row, product["total_col"])
        unit_coord = cell_coord(product_row, product["unit_col"])
        qty_coord = cell_coord(product_row, product["qty_col"])

        self._safe_set(ws, cell_coord(product_row, product["desc_col"]), product_description)
        self._safe_set(ws, qty_coord, quantity)
        self._safe_set(ws, unit_coord, unit_price)
        self._safe_set(ws, total_coord, f"={unit_coord}*{qty_coord}")
        self._safe_set(ws, product["unit_header"], f"Unit price({currency})")
        self._safe_set(ws, product["total_header"], f"Total Amount({currency})")

        if product.get("summary_total"):
            self._safe_set(ws, product["summary_total"], f"={total_coord}")
        if product.get("amount_words"):
            total = unit_price * quantity
            self._safe_set(ws, product["amount_words"], f"TOTAL AMOUNT: {currency} {total:.2f}")

    def _discover_layout(self, ws, *, template_path: Path, sheet_name: str) -> dict[str, Any]:
        context = f"{template_path.name}/{sheet_name}"

        invoice_date_label = self._find_label(ws, ("invoice date", "rechnungsdatum"))
        order_label = self._find_label(ws, ("order number", "bestellnummer"))
        invoice_label = self._find_exact_label(ws, ("invoice", "rechnung"))
        delivery_label = self._find_label(ws, ("delivery date", "lieferdatum"))
        customer_label = self._find_label(ws, ("customer information", "kundeninformationen", "to"))
        description_header = self._find_label(ws, ("description of goods", "beschreibung der waren"))
        unit_header = self._find_label(ws, ("unit price", "einzelpreis"))
        total_header = self._find_label(ws, ("total amount", "gesamtbetrag"))

        missing = [
            name
            for name, cell in [
                ("invoice date", invoice_date_label),
                ("order number", order_label),
                ("invoice number", invoice_label),
                ("delivery date", delivery_label),
                ("customer information", customer_label),
                ("description of goods", description_header),
                ("unit price", unit_header),
                ("total amount", total_header),
            ]
            if cell is None
        ]
        if missing:
            raise ValueError(f"Cannot identify {', '.join(missing)} in template sheet {context}")

        product_row = description_header.row + 1
        desc_col = description_header.column
        unit_col = unit_header.column
        total_col = total_header.column
        qty_col = self._find_quantity_col(ws, description_header, unit_col)

        return {
            "invoice_date": self._right_of(invoice_date_label),
            "order_number": self._right_of(order_label),
            "invoice_number": self._right_of(invoice_label),
            "delivery_date": self._right_of(delivery_label),
            "customer_info": self._customer_target(ws, customer_label),
            "product": {
                "row": product_row,
                "desc_col": desc_col,
                "qty_col": qty_col,
                "unit_col": unit_col,
                "total_col": total_col,
                "unit_header": unit_header.coordinate,
                "total_header": total_header.coordinate,
                "summary_total": self._summary_total_target(ws, product_row, total_col),
                "amount_words": self._amount_words_cell(ws),
            },
        }

    def _find_label(self, ws, terms: tuple[str, ...]):
        for row in ws.iter_rows():
            for cell in row:
                text = norm(cell.value)
                if text and any(term in text for term in terms):
                    return cell
        return None

    def _find_exact_label(self, ws, terms: tuple[str, ...]):
        for row in ws.iter_rows():
            for cell in row:
                text = norm(cell.value).rstrip(":")
                if text in terms:
                    return cell
        return None

    def _find_quantity_col(self, ws, description_header, unit_col: int) -> int:
        header_row = description_header.row
        for col in range(description_header.column + 1, unit_col):
            text = norm(ws.cell(header_row, col).value)
            if text in {"qty", "quantity", "menge"}:
                return col
        return max(description_header.column + 1, unit_col - 1)

    def _right_of(self, cell) -> str:
        return cell_coord(cell.row, cell.column + 1)

    def _customer_target(self, ws, label_cell) -> str:
        text = norm(label_cell.value).rstrip(":")
        right = ws.cell(label_cell.row, label_cell.column + 1)
        below_right = ws.cell(label_cell.row + 1, label_cell.column + 1)
        if text in {"to", "kundeninformationen"} and right.value not in (None, ""):
            return right.coordinate
        return below_right.coordinate

    def _summary_total_target(self, ws, product_row: int, total_col: int) -> str | None:
        for row in range(product_row + 1, min(ws.max_row, product_row + 5) + 1):
            row_text = [norm(ws.cell(row, col).value).rstrip(":") for col in range(1, min(ws.max_column, 5) + 1)]
            if "total" in row_text or "gesamt" in row_text:
                return cell_coord(row, total_col)
        return None

    def _amount_words_cell(self, ws) -> str | None:
        for row in ws.iter_rows():
            for cell in row:
                text = norm(cell.value)
                if text.startswith("total amount") or text.startswith("gesamtbetrag"):
                    return cell.coordinate
        return None

    def _safe_set(self, ws, coord: str, value: Any) -> None:
        cell = ws[coord]
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                anchor = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                ws[anchor].value = value
                return
        cell.value = value

    def export_pdf(self, xlsx_path: Path) -> Path:
        pdf_path = xlsx_path.with_suffix(".pdf")
        for cmd in ["libreoffice", "soffice"]:
            try:
                result = subprocess.run(
                    [
                        cmd,
                        "--headless",
                        "--convert-to",
                        "pdf",
                        "--outdir",
                        str(xlsx_path.parent),
                        str(xlsx_path),
                    ],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=30,
                )
                if result.returncode == 0 and pdf_path.exists():
                    return pdf_path
            except (FileNotFoundError, subprocess.TimeoutExpired):
                continue
        raise RuntimeError("LibreOffice not found. Install libreoffice or soffice to export PDF.")

    def _parse_date(self, value: Any) -> datetime:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        if isinstance(value, (int, float)):
            return datetime(1899, 12, 30) + timedelta(days=int(value))

        text = str(value).strip()
        for fmt in [
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d,%B,%Y",
            "%d,%b,%Y",
            "%d %B %Y",
            "%d %b %Y",
            "%d %B,%Y",
            "%d %b,%Y",
            "%B %d, %Y",
            "%b %d, %Y",
            "%d-%m-%Y",
            "%d/%m/%Y",
        ]:
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        raise ValueError(f"Cannot parse date: '{value}'")


def build_input_from_args(args: argparse.Namespace, parser: argparse.ArgumentParser) -> dict[str, Any]:
    if args.json:
        data = json.loads(args.json)
        if args.sheet and not data.get("sheet"):
            data["sheet"] = args.sheet
        if args.country and not data.get("country"):
            data["country"] = args.country
        if args.credit_note:
            data["credit_note"] = True
        return data

    missing = []
    for field in [
        "store",
        "order_number",
        "invoice_date",
        "delivery_date",
        "customer_info",
        "product_description",
    ]:
        if not getattr(args, field, None):
            missing.append(field)
    if not args.country and not args.sheet and not args.credit_note:
        missing.append("country_or_sheet")
    if args.quantity is None:
        missing.append("quantity")
    if args.unit_price is None:
        missing.append("unit_price")
    if missing:
        parser.error(f"Missing required fields: {', '.join(missing)}")

    return {
        "store": args.store,
        "country": args.country,
        "sheet": args.sheet,
        "order_number": args.order_number,
        "invoice_number": args.invoice_number or args.order_number,
        "invoice_date": args.invoice_date,
        "delivery_date": args.delivery_date,
        "customer_info": args.customer_info.replace("\\n", "\n"),
        "product_description": args.product_description,
        "quantity": args.quantity,
        "unit_price": args.unit_price,
        "credit_note": args.credit_note,
        "currency": args.currency,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Portable Commercial Invoice Generator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--store", help="Store/template name, e.g. 崔佳佳, 班威, ROB")
    parser.add_argument("--country", help="Backward-compatible sheet alias, e.g. 德国")
    parser.add_argument("--sheet", help="Exact worksheet name, e.g. 波兰, 德语发票, 贷记单")
    parser.add_argument("--order", dest="order_number", help="Order number")
    parser.add_argument("--invoice-num", dest="invoice_number", help="Invoice number")
    parser.add_argument("--date", dest="invoice_date", help="Invoice date")
    parser.add_argument("--delivery", dest="delivery_date", help="Delivery date")
    parser.add_argument("--customer", dest="customer_info", help="Buyer info, use \\n for line breaks")
    parser.add_argument("--product", dest="product_description", help="Product description")
    parser.add_argument("--qty", dest="quantity", type=int, help="Quantity")
    parser.add_argument("--price", dest="unit_price", type=float, help="Unit price")
    parser.add_argument("--currency", default="EUR", help="Currency, default EUR")
    parser.add_argument("--credit-note", action="store_true", help="Generate credit note")
    parser.add_argument("--pdf", action="store_true", help="Export PDF with LibreOffice/soffice")
    parser.add_argument("--list-stores", action="store_true", help="List available stores")
    parser.add_argument("--list-countries", action="store_true", help="List country aliases")
    parser.add_argument("--list-sheets", action="store_true", help="List business sheets for --store")
    parser.add_argument("--json", help="JSON input using script field names")
    args = parser.parse_args()

    generator = InvoiceGenerator()

    if args.list_stores:
        print("Available stores:")
        for store in generator.list_stores():
            print(f"  - {store}")
        return

    if args.list_countries:
        print("Supported country aliases:")
        for country in generator.list_countries():
            print(f"  - {country}")
        return

    if args.list_sheets:
        if not args.store:
            parser.error("--list-sheets requires --store")
        print(f"Available sheets for {args.store}:")
        for sheet in generator.list_sheets(args.store):
            print(f"  - {sheet}")
        return

    data = build_input_from_args(args, parser)
    result = generator.generate(**data)
    print(f"Generated: {result}")

    if args.pdf:
        try:
            pdf = generator.export_pdf(result)
            print(f"PDF exported: {pdf}")
        except RuntimeError as exc:
            print(f"PDF export skipped: {exc}")


if __name__ == "__main__":
    main()
